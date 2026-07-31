#!/usr/bin/env python3
"""Read-only structural audit for standard Yayoi Accounting Excel exports.

The script never connects to a database and never prints accounting amounts.
It emits workbook structure, labels, formula counts, and cell-type metadata as
JSON so that parser feasibility can be reviewed without copying financial data
into logs or test fixtures.
"""

from __future__ import annotations

import argparse
import hashlib
import json
import re
import sys
from collections import Counter
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.utils import get_column_letter


STATEMENT_HINTS = {
    "bs": ("貸借対照表", "資産の部", "負債の部", "純資産の部"),
    "pl": ("損益計算書", "売上高", "売上総利益", "営業利益", "経常利益"),
}
PERIOD_RE = re.compile(r"(?:20\d{2}|令和\d+)年|\d{1,2}月|決算|累計|期間")
ACCOUNT_CODE_RE = re.compile(r"^\d{3,8}$")


def normalized_text(value: Any) -> str | None:
    if not isinstance(value, str):
        return None
    text = " ".join(value.replace("\u3000", " ").split())
    return text or None


def redact_path(path: Path) -> str:
    return path.name


def workbook_hash(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as source:
        for chunk in iter(lambda: source.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def classify_statement(labels: list[str]) -> str:
    scores = {
        key: sum(1 for hint in hints if any(hint in label for label in labels))
        for key, hints in STATEMENT_HINTS.items()
    }
    if scores["bs"] > scores["pl"]:
        return "BS"
    if scores["pl"] > scores["bs"]:
        return "PL"
    return "Unknown"


def numeric_value(sheet: Any, row: int, column: int) -> float | None:
    value = sheet.cell(row=row, column=column).value
    if isinstance(value, bool) or not isinstance(value, (int, float)):
        return None
    return float(value)


def arithmetic_check(
    sheet: Any,
    name: str,
    expected_row: int,
    component_rows: list[tuple[int, float]],
) -> dict[str, Any]:
    checked = 0
    failed_cells: list[str] = []
    skipped_cells: list[str] = []
    for column in range(2, 19):
        expected = numeric_value(sheet, expected_row, column)
        components = [
            numeric_value(sheet, row, column) for row, _ in component_rows
        ]
        coordinate = f"{get_column_letter(column)}{expected_row}"
        if expected is None or any(value is None for value in components):
            skipped_cells.append(coordinate)
            continue
        calculated = sum(
            value * multiplier
            for value, (_, multiplier) in zip(components, component_rows)
        )
        checked += 1
        if abs(expected - calculated) > 0.5:
            failed_cells.append(coordinate)
    return {
        "name": name,
        "checked_cell_count": checked,
        "passed": checked > 0 and not failed_cells,
        "failed_cells": failed_cells,
        "skipped_cells": skipped_cells,
    }


def sheet_checks(sheet: Any, statement_type: str) -> list[dict[str, Any]]:
    if statement_type == "BS":
        balance_row = 130 if sheet.max_row == 130 else 131
        return [
            arithmetic_check(
                sheet,
                "assets_equal_liabilities_and_equity",
                76,
                [(balance_row, 1.0)],
            )
        ]
    if statement_type == "PL":
        return [
            arithmetic_check(
                sheet,
                "gross_profit",
                24,
                [(13, 1.0), (23, -1.0)],
            ),
            arithmetic_check(
                sheet,
                "operating_profit",
                57,
                [(24, 1.0), (56, -1.0)],
            ),
            arithmetic_check(
                sheet,
                "ordinary_profit",
                74,
                [(57, 1.0), (65, 1.0), (73, -1.0)],
            ),
            arithmetic_check(
                sheet,
                "pretax_profit",
                85,
                [(74, 1.0), (78, 1.0), (83, -1.0)],
            ),
            arithmetic_check(
                sheet,
                "net_profit",
                87,
                [(85, 1.0), (86, -1.0)],
            ),
        ]
    return []


def cross_sheet_check(
    workbook: Any,
    name: str,
    target_name: str,
    component_names: list[str],
) -> dict[str, Any]:
    if target_name not in workbook.sheetnames or any(
        item not in workbook.sheetnames for item in component_names
    ):
        return {
            "name": name,
            "status": "not_applicable",
            "target_sheet": target_name,
            "component_sheet_count": len(component_names),
        }
    target = workbook[target_name]
    components = [workbook[item] for item in component_names]
    failed_cells: list[str] = []
    checked = 0
    for row in range(9, target.max_row + 1):
        for column in range(2, 19):
            target_value = numeric_value(target, row, column)
            values = [numeric_value(sheet, row, column) for sheet in components]
            if target_value is None or any(value is None for value in values):
                continue
            checked += 1
            if abs(target_value - sum(values)) > 0.5:
                failed_cells.append(f"{get_column_letter(column)}{row}")
    return {
        "name": name,
        "status": "pass" if checked and not failed_cells else "fail",
        "target_sheet": target_name,
        "component_sheet_count": len(component_names),
        "checked_cell_count": checked,
        "failed_cell_count": len(failed_cells),
        "failed_cells_sample": failed_cells[:40],
    }


def workbook_reconciliations(workbook: Any) -> list[dict[str, Any]]:
    checks: list[dict[str, Any]] = []
    for prefix in ("貸･", "損･"):
        checks.append(
            cross_sheet_check(
                workbook,
                f"{prefix}head_office_rollup",
                f"{prefix}本部",
                [
                    f"{prefix}本部･営業",
                    f"{prefix}本部･教育(合計)",
                    f"{prefix}本部･総務",
                    f"{prefix}本部･経理",
                ],
            )
        )
        checks.append(
            cross_sheet_check(
                workbook,
                f"{prefix}education_rollup",
                f"{prefix}本部･教育(合計)",
                [f"{prefix}教育･アカデミー", f"{prefix}本部･教育(共通)"],
            )
        )
        fc_components = [
            name
            for name in workbook.sheetnames
            if name.startswith(f"{prefix}FC")
            and name not in {f"{prefix}FC(合計)"}
        ]
        checks.append(
            cross_sheet_check(
                workbook,
                f"{prefix}fc_rollup",
                f"{prefix}FC(合計)",
                fc_components,
            )
        )
        total_components = [
            f"{prefix}本部",
            f"{prefix}KYARA HALF",
            *[
                name
                for name in workbook.sheetnames
                if name.startswith(f"{prefix}BASSA")
            ],
            f"{prefix}EC事業部",
            f"{prefix}FC(合計)",
            f"{prefix}全体(共通)",
        ]
        checks.append(
            cross_sheet_check(
                workbook,
                f"{prefix}all_entity_rollup",
                f"{prefix}全体(合計)",
                total_components,
            )
        )
    return checks


def analyze_sheet(sheet: Any) -> dict[str, Any]:
    type_counts: Counter[str] = Counter()
    formula_cells: list[str] = []
    labels_with_cells: list[tuple[str, str, dict[str, Any]]] = []
    period_labels: list[dict[str, str]] = []
    possible_account_codes: list[dict[str, str]] = []
    numeric_columns: Counter[str] = Counter()
    nonempty_by_row: Counter[int] = Counter()
    nonempty_by_column: Counter[int] = Counter()

    for row in sheet.iter_rows():
        for cell in row:
            if isinstance(cell, MergedCell):
                continue
            value = cell.value
            if value is None:
                type_counts["blank"] += 1
                continue
            nonempty_by_row[cell.row] += 1
            nonempty_by_column[cell.column] += 1
            if cell.data_type == "f":
                type_counts["formula"] += 1
                formula_cells.append(cell.coordinate)
                continue
            if isinstance(value, bool):
                type_counts["boolean"] += 1
            elif isinstance(value, (int, float)):
                type_counts["numeric"] += 1
                numeric_columns[get_column_letter(cell.column)] += 1
            else:
                text = normalized_text(value)
                if not text:
                    type_counts["blank_text"] += 1
                    continue
                type_counts["text"] += 1
                labels_with_cells.append(
                    (
                        cell.coordinate,
                        text,
                        {
                            "style_id": cell.style_id,
                            "indent": cell.alignment.indent or 0,
                            "bold": bool(cell.font.bold),
                        },
                    )
                )
                if PERIOD_RE.search(text):
                    period_labels.append({"cell": cell.coordinate, "label": text})
                if ACCOUNT_CODE_RE.match(text):
                    possible_account_codes.append(
                        {"cell": cell.coordinate, "code": text}
                    )

    labels = [label for _, label, _ in labels_with_cells]
    statement_type = (
        "BS"
        if sheet.title.startswith("貸･")
        else "PL"
        if sheet.title.startswith("損･")
        else classify_statement(labels)
    )
    keyword_labels = [
        {"cell": cell, "label": label}
        for cell, label, _ in labels_with_cells
        if any(
            keyword in label
            for keyword in (
                "貸借対照表",
                "損益計算書",
                "残高",
                "月度",
                "累計",
                "部門",
                "全社",
                "売上",
                "利益",
                "資産",
                "負債",
                "純資産",
                "費",
                "仕入",
            )
        )
    ]

    likely_header_rows = [
        {"row": row_no, "nonempty_cells": count}
        for row_no, count in nonempty_by_row.items()
        if row_no <= 20 and count >= 2
    ]
    numeric_column_summary = [
        {
            "column": get_column_letter(column),
            "numeric_cells": sum(
                1
                for row in range(9, sheet.max_row + 1)
                if numeric_value(sheet, row, column) is not None
            ),
            "positive_cells": sum(
                1
                for row in range(9, sheet.max_row + 1)
                if (numeric_value(sheet, row, column) or 0) > 0
            ),
            "negative_cells": sum(
                1
                for row in range(9, sheet.max_row + 1)
                if (numeric_value(sheet, row, column) or 0) < 0
            ),
            "zero_cells": sum(
                1
                for row in range(9, sheet.max_row + 1)
                if numeric_value(sheet, row, column) == 0
            ),
            "blank_cells": sum(
                1
                for row in range(9, sheet.max_row + 1)
                if sheet.cell(row=row, column=column).value is None
            ),
        }
        for column in range(2, 19)
    ]

    return {
        "sheet_name": sheet.title,
        "sheet_state": sheet.sheet_state,
        "statement_type": statement_type,
        "used_range": sheet.calculate_dimension(),
        "max_row": sheet.max_row,
        "max_column": sheet.max_column,
        "merged_ranges": [str(item) for item in sheet.merged_cells.ranges],
        "cell_type_counts": dict(type_counts),
        "formula_cell_count": len(formula_cells),
        "formula_cells_sample": formula_cells[:20],
        "likely_header_rows": likely_header_rows,
        "period_labels": period_labels[:80],
        "possible_account_codes": possible_account_codes[:80],
        "numeric_columns": numeric_column_summary,
        "june_equals_july_for_all_numeric_cells": all(
            numeric_value(sheet, row, 12) == numeric_value(sheet, row, 13)
            for row in range(9, sheet.max_row + 1)
            if numeric_value(sheet, row, 12) is not None
            and numeric_value(sheet, row, 13) is not None
        ),
        "keyword_labels": keyword_labels[:500],
        "all_text_labels": [
            {"cell": cell, "label": label, **style}
            for cell, label, style in labels_with_cells
        ],
        "arithmetic_checks": sheet_checks(sheet, statement_type),
    }


def main() -> int:
    parser = argparse.ArgumentParser(
        description="Audit a Yayoi Excel export without emitting accounting amounts."
    )
    parser.add_argument("input", type=Path, help="Path to the original .xlsx file")
    parser.add_argument(
        "--output",
        type=Path,
        help="Optional JSON output path. Defaults to stdout.",
    )
    args = parser.parse_args()

    source = args.input.resolve()
    if not source.is_file():
        parser.error(f"input file not found: {source}")
    if source.suffix.lower() != ".xlsx":
        parser.error("input must be an .xlsx file")

    try:
        workbook = load_workbook(
            filename=source,
            read_only=False,
            data_only=False,
            keep_links=False,
        )
    except Exception as exc:
        print(f"failed to read workbook: {type(exc).__name__}", file=sys.stderr)
        return 2

    result = {
        "source_file": redact_path(source),
        "sha256": workbook_hash(source),
        "sheet_count": len(workbook.sheetnames),
        "defined_name_count": len(workbook.defined_names),
        "sheets": [analyze_sheet(workbook[name]) for name in workbook.sheetnames],
        "workbook_reconciliations": workbook_reconciliations(workbook),
    }
    payload = json.dumps(result, ensure_ascii=False, indent=2)

    if args.output:
        args.output.parent.mkdir(parents=True, exist_ok=True)
        args.output.write_text(payload + "\n", encoding="utf-8")
        print(
            json.dumps(
                {
                    "status": "ok",
                    "sheet_count": result["sheet_count"],
                    "output": str(args.output),
                },
                ensure_ascii=False,
            )
        )
    else:
        print(payload)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
