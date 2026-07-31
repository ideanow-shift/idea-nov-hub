from __future__ import annotations

import hashlib
import re
from collections import Counter
from datetime import date
from decimal import Decimal
from pathlib import Path
from typing import Iterator

from openpyxl import load_workbook
from openpyxl.cell.cell import TYPE_ERROR

from .domain import RawValue, ScopeType, StatementType, ValueState

MONTH_RE = re.compile(r"(?:(\d{4})年)?(\d{1,2})月")
FISCAL_START_RE = re.compile(r"(?:令和(\d{1,2})年|((?:19|20)\d{2})年)(\d{1,2})月(\d{1,2})日")


def _anchor(value: object) -> str:
    return str(value or "").replace("：", ":").replace("（", "(").replace("）", ")").replace(" ", "")


def file_hash(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as source:
        for chunk in iter(lambda: source.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _scope(entity: str) -> ScopeType:
    if entity == "全体(合計)":
        return ScopeType.LEGAL_ENTITY_SUMMARY
    if entity == "FC(合計)":
        return ScopeType.FRANCHISE_SUMMARY
    if "(合計)" in entity or entity.startswith("本部"):
        return ScopeType.DEPARTMENT_SUMMARY
    if entity.startswith("FC"):
        return ScopeType.LEAF_STORE
    if entity.startswith(("BASSA", "KYARA")):
        return ScopeType.LEAF_STORE
    return ScopeType.UNKNOWN


def _state(cell, cached_value) -> tuple[ValueState, Decimal | None, str | None]:
    if cell.data_type == "f":
        numeric = cached_value if isinstance(cached_value, (int, float)) else None
        return ValueState.FORMULA, Decimal(str(numeric)) if numeric is not None else None, str(cell.value)
    if cell.data_type == TYPE_ERROR:
        return ValueState.ERROR, None, None
    value = cell.value
    if value is None:
        return ValueState.BLANK, None, None
    if isinstance(value, bool):
        return ValueState.TEXT, None, None
    if isinstance(value, (int, float)):
        return (ValueState.ZERO if value == 0 else ValueState.AMOUNT), Decimal(str(value)), None
    return ValueState.TEXT, None, None


def _fiscal_start_year(period_label: object) -> int:
    match = FISCAL_START_RE.search(str(period_label or ""))
    if not match:
        raise ValueError("fiscal period start is not recognizable")
    return 2018 + int(match.group(1)) if match.group(1) else int(match.group(2))


def _period(label: str, fiscal_start_year: int) -> date | None:
    match = MONTH_RE.search(label)
    if not match:
        return None
    year = int(match.group(1)) if match.group(1) else fiscal_start_year
    month = int(match.group(2))
    if not match.group(1) and month < 9:
        year += 1
    return date(year, month, 1)


class YayoiExcelAdapter:
    source_system = "yayoi_excel"

    def __init__(self, path: Path):
        self.path = path
        # Normal mode is materially faster for repeated A:R access. The adapter
        # never calls save(), so the source file remains read-only in practice.
        self.workbook = load_workbook(path, read_only=False, data_only=False)
        self.cached = load_workbook(path, read_only=False, data_only=True)

    def close(self) -> None:
        self.workbook.close()
        self.cached.close()

    def extract(self) -> Iterator[RawValue]:
        for sheet_name in self.workbook.sheetnames:
            sheet = self.workbook[sheet_name]
            cached_sheet = self.cached[sheet_name]
            report_anchor = _anchor(sheet["A1"].value)
            if "残高試算表(年間推移)" not in report_anchor or _anchor(sheet["A8"].value) != "勘定科目":
                raise ValueError(f"major layout drift: {sheet_name}")
            tax_anchor = _anchor(sheet["A6"].value)
            if "税抜" not in tax_anchor or tax_anchor.endswith("税込"):
                raise ValueError(f"unsupported tax basis: {sheet_name}")
            fiscal_start_year = _fiscal_start_year(sheet["A5"].value)
            prefix, _, entity = sheet_name.partition("･")
            statement = StatementType.BS if prefix == "貸" else StatementType.PL if prefix == "損" else None
            if statement is None or not entity:
                raise ValueError(f"unknown sheet type: {sheet_name}")
            section = None
            occurrences: Counter[str] = Counter()
            headers = {column: str(sheet.cell(8, column).value or "") for column in range(2, 19)}
            for row in range(9, sheet.max_row + 1):
                label = str(sheet.cell(row, 1).value or "").strip()
                if not label:
                    continue
                occurrences[label] += 1
                row_has_number = any(
                    isinstance(sheet.cell(row, column).value, (int, float))
                    for column in range(2, 19)
                )
                if not row_has_number:
                    section = label
                    continue
                occurrence = f"{section or 'root'}:{label}:{occurrences[label]}"
                for column in range(2, 19):
                    cell = sheet.cell(row, column)
                    cached_cell = cached_sheet.cell(row, column)
                    state, amount, formula = _state(cell, cached_cell.value)
                    yield RawValue(
                        source_sheet=sheet_name,
                        source_row=row,
                        source_column=column,
                        source_column_label=headers[column],
                        fiscal_year=fiscal_start_year,
                        detected_period=_period(headers[column], fiscal_start_year),
                        statement_type=statement,
                        source_entity_name=entity,
                        scope_type=_scope(entity),
                        section=section,
                        source_account_name=label,
                        parent_context=section,
                        occurrence_context=occurrence,
                        value_state=state,
                        amount_net=amount,
                        formula=formula,
                    )
