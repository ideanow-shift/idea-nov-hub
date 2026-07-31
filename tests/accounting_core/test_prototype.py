from __future__ import annotations

import sqlite3
import tempfile
import unittest
from datetime import date
from pathlib import Path

from openpyxl import Workbook

from accounting_core.domain import MappingStatus, ValueState
from accounting_core.mapping import CsvMappingReader
from accounting_core.normalization import normalize
from accounting_core.staging import open_database
from accounting_core.validation import validate_mappings, validate_raw
from accounting_core.yayoi_excel import YayoiExcelAdapter


def synthetic_workbook(path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "損･TEST店"
    sheet["A1"] = "残高試算表（年間推移）"
    sheet["A3"] = "Synthetic Corporation"
    sheet["A5"] = "2025年9月1日～2026年8月31日"
    sheet["A6"] = "税抜"
    sheet["A8"] = "勘定科目"
    headers = ["9月", "10月", "11月", "12月", "1月", "2月", "上半期", "3月", "4月", "5月", "6月", "7月", "8月", "下半期", "当期仮残高", "決算残高", "当期残高"]
    for column, label in enumerate(headers, 2):
        sheet.cell(8, column, label)
    labels = {
        9: ("[売上高]", None),
        10: ("売上高合計", 100),
        11: ("[売上原価]", None),
        12: ("売上原価", 40),
        13: ("売上総損益金額", 60),
        14: ("[販売管理費]", None),
        15: ("販売管理費計", 30),
        16: ("営業損益金額", 30),
        17: ("営業外収益合計", 2),
        18: ("営業外費用合計", 1),
        19: ("経常損益金額", 31),
        20: ("特別利益合計", 0),
        21: ("特別損失合計", 0),
        22: ("税引前当期純損益金額", 31),
        23: ("法人税及住民税", 5),
        24: ("当期純損益金額", 26),
    }
    for row, (label, amount) in labels.items():
        sheet.cell(row, 1, label)
        if amount is not None:
            for column in range(2, 19):
                sheet.cell(row, column, amount)
    workbook.save(path)


class PrototypeTests(unittest.TestCase):
    def test_adapter_distinguishes_zero_blank_and_period(self):
        with tempfile.TemporaryDirectory() as directory:
            path = Path(directory) / "synthetic.xlsx"
            synthetic_workbook(path)
            adapter = YayoiExcelAdapter(path)
            values = list(adapter.extract())
            adapter.close()
        self.assertTrue(any(v.value_state is ValueState.ZERO for v in values))
        self.assertTrue(any(v.detected_period and v.detected_period.isoformat() == "2026-06-01" for v in values))
        self.assertTrue(all(v.source_entity_name == "TEST店" for v in values))

    def test_validation_accepts_equations_and_blocks_unconfirmed_period(self):
        with tempfile.TemporaryDirectory() as directory:
            path = Path(directory) / "synthetic.xlsx"
            synthetic_workbook(path)
            adapter = YayoiExcelAdapter(path)
            results = validate_raw(adapter.extract())
            adapter.close()
        codes = {result.code for result in results}
        self.assertNotIn("PL_GROSS_PROFIT_MISMATCH", codes)
        self.assertIn("PERIOD_DUPLICATE_CAUSE_UNKNOWN", codes)

    def test_confirmed_cutoff_classifies_only_later_duplicate_as_carry_forward(self):
        with tempfile.TemporaryDirectory() as directory:
            path = Path(directory) / "synthetic.xlsx"
            synthetic_workbook(path)
            adapter = YayoiExcelAdapter(path)
            values = list(adapter.extract())
            adapter.close()
        results = validate_raw(values, date(2026, 6, 1))
        carry = [result for result in results if result.code == "UNCONFIRMED_PERIOD_CARRY_FORWARD"]
        self.assertEqual(1, len(carry))
        self.assertEqual("pending", carry[0].closing_status)
        self.assertFalse(carry[0].publish_allowed)
        self.assertEqual("preparing", carry[0].data_state)

    def test_cutoff_controls_canonical_period_state(self):
        root = Path(__file__).parents[2]
        reader = CsvMappingReader(
            root / "docs/accounting/yayoi-entity-mapping.csv",
            root / "docs/accounting/yayoi-account-mapping.csv",
        )
        with tempfile.TemporaryDirectory() as directory:
            path = Path(directory) / "synthetic.xlsx"
            synthetic_workbook(path)
            adapter = YayoiExcelAdapter(path)
            values = list(adapter.extract())
            adapter.close()
        june = next(value for value in values if value.detected_period == date(2026, 6, 1))
        july = next(value for value in values if value.detected_period == date(2026, 7, 1))
        june_fact = normalize(june, reader, date(2026, 6, 1))
        july_fact = normalize(july, reader, date(2026, 6, 1))
        self.assertEqual(("confirmed", "available", True), (june_fact.closing_status, june_fact.data_state, june_fact.publish_allowed))
        self.assertEqual(("pending", "preparing", False), (july_fact.closing_status, july_fact.data_state, july_fact.publish_allowed))

    def test_duplicate_inside_confirmed_period_is_not_auto_carry_forward(self):
        with tempfile.TemporaryDirectory() as directory:
            path = Path(directory) / "synthetic.xlsx"
            synthetic_workbook(path)
            adapter = YayoiExcelAdapter(path)
            values = list(adapter.extract())
            adapter.close()
        codes = {result.code for result in validate_raw(values, date(2026, 7, 1))}
        self.assertIn("PERIOD_DUPLICATE_CAUSE_UNKNOWN", codes)
        self.assertNotIn("UNCONFIRMED_PERIOD_CARRY_FORWARD", codes)

    def test_unapproved_mappings_block_normalization(self):
        root = Path(__file__).parents[2]
        reader = CsvMappingReader(
            root / "docs/accounting/yayoi-entity-mapping.csv",
            root / "docs/accounting/yayoi-account-mapping.csv",
        )
        with tempfile.TemporaryDirectory() as directory:
            path = Path(directory) / "synthetic.xlsx"
            synthetic_workbook(path)
            adapter = YayoiExcelAdapter(path)
            facts = [normalize(raw, reader) for raw in adapter.extract()]
            adapter.close()
        self.assertTrue(all(not fact.publishable for fact in facts))
        codes = {r.code for r in validate_mappings(facts)}
        self.assertIn("ENTITY_MAPPING_NOT_APPROVED", codes)
        self.assertNotIn("DUPLICATE_FACT", codes)

    def test_historical_account_alias_remains_proposed(self):
        root = Path(__file__).parents[2]
        reader = CsvMappingReader(
            root / "docs/accounting/yayoi-entity-mapping.csv",
            root / "docs/accounting/yayoi-account-mapping.csv",
        )
        with tempfile.TemporaryDirectory() as directory:
            path = Path(directory) / "synthetic.xlsx"
            synthetic_workbook(path)
            adapter = YayoiExcelAdapter(path)
            raw = next(value for value in adapter.extract() if value.source_account_name == "売上高合計")
            adapter.close()
        historical = raw.__class__(**{**raw.__dict__, "source_account_name": "売上高", "fiscal_year": 2024})
        resolution = reader.resolve_account(historical)
        self.assertEqual("total_revenue", resolution.target)
        self.assertEqual(MappingStatus.PROPOSED, resolution.status)
        future = raw.__class__(**{**raw.__dict__, "source_account_name": "売上高", "fiscal_year": 2026})
        self.assertIsNone(reader.resolve_account(future).target)

    def test_schema_has_required_tables_and_published_facts_are_immutable(self):
        connection = open_database()
        tables = {
            row[0] for row in connection.execute(
                "SELECT name FROM sqlite_master WHERE type='table' AND name LIKE 'accounting_%'"
            )
        }
        self.assertEqual(11, len(tables))
        connection.execute("INSERT INTO accounting_import_batches(id,source_system,status,created_by) VALUES('b','test','imported','actor')")
        connection.execute("INSERT INTO accounting_import_files(id,batch_id,source_system,file_hash,original_file_name,confirmed_through_period) VALUES('f','b','test','h','masked.xlsx','2026-06-01')")
        connection.execute("""INSERT INTO accounting_versions(
          id,version_number,version_label,fiscal_year,fiscal_month,version_type,
          scope_type,scope_id,import_file_id,status,created_by
        ) VALUES('v',1,'2026-06-FINAL-01',2026,6,'final','store','e','f','published','actor')""")
        connection.execute("""INSERT INTO accounting_raw_values(
          id,import_file_id,source_sheet,source_sheet_type,source_row,source_column,
          source_cell_reference,source_column_label,fiscal_year,source_value_state,statement_type,
          source_entity_name,scope_type,source_account_name
        ) VALUES('r','f','sheet','pl',1,2,'B1','6月',2025,'amount','pl','entity','leaf_store','sales')""")
        connection.execute("""INSERT INTO accounting_facts(
          id,raw_value_id,version_id,source_file_id,normalized_account,entity_id,
          scope_type,scope_id,period,amount_net,tax_basis,status
        ) VALUES('x','r','v','f','sales','e','store','e','2026-06-01','1','tax_exclusive','published')""")
        with self.assertRaises(sqlite3.IntegrityError):
            connection.execute("UPDATE accounting_facts SET amount_net='2' WHERE id='x'")


if __name__ == "__main__":
    unittest.main()
